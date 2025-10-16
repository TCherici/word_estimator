#!/usr/bin/env python3
# /// script
# requires-python = ">=3.8"
# dependencies = [
#     "PyQt5>=5.15.10",
#     "PyMuPDF>=1.23.8",
#     "pytesseract>=0.3.10",
#     "pdf2image>=1.16.3",
#     "Pillow>=10.2.0",
#     "openpyxl>=3.1.2",
# ]
# ///
"""
PDF Keyword Value Estimator
A GUI application to count keyword occurrences in PDF files and calculate their total value.
"""

import sys
import re
import csv
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime

import fitz  # PyMuPDF
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QFileDialog, QTextEdit, QMessageBox, QHeaderView, QGroupBox,
    QProgressDialog, QRadioButton, QButtonGroup
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont


class OCRWorker(QThread):
    """Worker thread for OCR processing to avoid blocking the UI."""

    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    progress = pyqtSignal(int, str)

    def __init__(self, pdf_path):
        super().__init__()
        self.pdf_path = pdf_path

    def run(self):
        """Run OCR processing in a separate thread."""
        try:
            # First, determine the number of pages
            self.progress.emit(0, "Opening PDF and counting pages...")
            doc = fitz.open(self.pdf_path)
            total_pages = doc.page_count
            doc.close()

            if total_pages == 0:
                self.error.emit("No pages found in PDF")
                return

            self.progress.emit(5, f"Found {total_pages} page(s). Starting conversion...")

            text = ""

            # Process pages one at a time to show accurate progress
            for page_num in range(1, total_pages + 1):
                # Calculate progress (5-95% for conversion and OCR, OCR being the bulk of the work)
                # Give 5% initial + 5% conversion/setup + 90% OCR processing per page
                progress_percent = 5 + int(((page_num - 1) / total_pages) * 90)

                # Convert single page to image (quick step, ~10% of time per page)
                conversion_progress = progress_percent + int((1 / total_pages) * 9)
                self.progress.emit(
                    conversion_progress,
                    f"Converting page {page_num} of {total_pages} to image..."
                )

                # Convert this specific page
                images = convert_from_path(
                    self.pdf_path,
                    dpi=300,
                    first_page=page_num,
                    last_page=page_num
                )

                if len(images) == 0:
                    continue

                image = images[0]

                # Perform OCR on the image (slow step, ~90% of time per page)
                ocr_progress = 5 + int((page_num / total_pages) * 90)
                self.progress.emit(
                    ocr_progress,
                    f"Extracting text from page {page_num} of {total_pages}..."
                )

                page_text = pytesseract.image_to_string(image)
                text += page_text + "\n"

            self.progress.emit(100, "OCR complete!")
            self.finished.emit(text)

        except Exception as e:
            self.error.emit(f"OCR Error: {str(e)}")


class KeywordEstimatorApp(QMainWindow):
    """Main application window for PDF keyword value estimation."""

    def __init__(self):
        super().__init__()
        self.pdf_path = None
        self.ocr_worker = None
        self.progress_dialog = None
        self.keywords_file = Path.home() / ".word_estimator" / "keywords.csv"
        self.current_results = None  # Store results for Excel export
        self.init_ui()
        self.load_keywords()

    def get_keywords_dir(self) -> Path:
        """Get or create the keywords directory."""
        keywords_dir = Path.home() / ".word_estimator"
        keywords_dir.mkdir(parents=True, exist_ok=True)
        return keywords_dir

    def load_keywords(self):
        """Load keywords from CSV file and populate the table."""
        try:
            if not self.keywords_file.exists():
                return

            with open(self.keywords_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                keywords = list(reader)

            if not keywords:
                return

            # Clear existing table
            self.keywords_table.clearContents()
            self.keywords_table.setRowCount(len(keywords))

            # Populate table with loaded keywords
            for row, (keyword, value) in enumerate(keywords):
                self.keywords_table.setItem(row, 0, QTableWidgetItem(keyword))
                self.keywords_table.setItem(row, 1, QTableWidgetItem(value))

        except Exception as e:
            QMessageBox.warning(
                self,
                "Load Error",
                f"Could not load keywords from file:\n{str(e)}\n\nStarting with empty table."
            )

    def save_keywords(self):
        """Save keywords from table to CSV file (silent mode for auto-save)."""
        try:
            # Ensure directory exists
            self.get_keywords_dir()

            # Get all keywords from table
            keywords = []
            for row in range(self.keywords_table.rowCount()):
                keyword_item = self.keywords_table.item(row, 0)
                value_item = self.keywords_table.item(row, 1)

                if keyword_item and value_item:
                    keyword = keyword_item.text().strip()
                    value = value_item.text().strip()

                    if keyword and value:
                        keywords.append([keyword, value])

            # Write to CSV
            with open(self.keywords_file, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(keywords)

        except Exception as e:
            QMessageBox.critical(
                self,
                "Save Error",
                f"Could not save keywords to file:\n{str(e)}"
            )

    def manual_save_keywords(self):
        """Manually save keywords with user confirmation."""
        self.save_keywords()
        QMessageBox.information(
            self,
            "Keywords Saved",
            f"Keywords have been saved to:\n{self.keywords_file}"
        )

    def init_ui(self):
        """Initialize the user interface."""
        self.setWindowTitle("PDF Keyword Value Estimator")
        self.showMaximized()  # Start in fullscreen/maximized mode

        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)

        # PDF File Selection Section
        file_group = QGroupBox("1. Select PDF File")
        file_layout = QHBoxLayout()

        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("color: #666;")
        file_layout.addWidget(self.file_label, stretch=1)

        self.browse_btn = QPushButton("Browse...")
        self.browse_btn.clicked.connect(self.browse_pdf)
        self.browse_btn.setMinimumWidth(100)
        file_layout.addWidget(self.browse_btn)

        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # Keywords Input Section
        keywords_group = QGroupBox("2. Enter Keywords and Values")
        keywords_layout = QVBoxLayout()

        # Table for keywords and values
        self.keywords_table = QTableWidget()
        self.keywords_table.setColumnCount(2)
        self.keywords_table.setHorizontalHeaderLabels(["Keyword", "Value (Integer)"])
        self.keywords_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.keywords_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.keywords_table.setRowCount(5)  # Start with 5 empty rows
        keywords_layout.addWidget(self.keywords_table)

        # Buttons for table management
        table_buttons_layout = QHBoxLayout()

        self.add_row_btn = QPushButton("Add Row")
        self.add_row_btn.clicked.connect(self.add_table_row)
        table_buttons_layout.addWidget(self.add_row_btn)

        self.remove_row_btn = QPushButton("Remove Selected Row")
        self.remove_row_btn.clicked.connect(self.remove_table_row)
        table_buttons_layout.addWidget(self.remove_row_btn)

        self.clear_table_btn = QPushButton("Clear All")
        self.clear_table_btn.clicked.connect(self.clear_table)
        table_buttons_layout.addWidget(self.clear_table_btn)

        table_buttons_layout.addStretch()

        self.save_keywords_btn = QPushButton("Save Keywords")
        self.save_keywords_btn.clicked.connect(self.manual_save_keywords)
        table_buttons_layout.addWidget(self.save_keywords_btn)

        self.load_keywords_btn = QPushButton("Reload Keywords")
        self.load_keywords_btn.clicked.connect(self.load_keywords)
        table_buttons_layout.addWidget(self.load_keywords_btn)

        keywords_layout.addLayout(table_buttons_layout)

        keywords_group.setLayout(keywords_layout)
        main_layout.addWidget(keywords_group)

        # Calculate Button
        self.calculate_btn = QPushButton("Calculate Total")
        self.calculate_btn.clicked.connect(self.calculate_total)
        self.calculate_btn.setMinimumHeight(40)
        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        self.calculate_btn.setFont(font)
        self.calculate_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        main_layout.addWidget(self.calculate_btn)

        # Extracted Text Section
        extracted_group = QGroupBox("3. Extracted Text Preview")
        extracted_layout = QVBoxLayout()

        self.extracted_text = QTextEdit()
        self.extracted_text.setReadOnly(True)
        self.extracted_text.setMinimumHeight(250)
        self.extracted_text.setPlaceholderText("Extracted text from PDF will appear here after processing...")
        font = QFont("Arial")
        font.setPointSize(9)
        self.extracted_text.setFont(font)
        extracted_layout.addWidget(self.extracted_text)

        extracted_group.setLayout(extracted_layout)
        main_layout.addWidget(extracted_group)

        # Results Section
        results_group = QGroupBox("4. Calculation Results")
        results_layout = QVBoxLayout()

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(4)
        self.results_table.setHorizontalHeaderLabels(["Keyword", "Count", "Value", "Subtotal"])
        self.results_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.results_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.results_table.setMinimumHeight(200)
        self.results_table.setSortingEnabled(True)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)  # Read-only
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        results_layout.addWidget(self.results_table)

        # Export button
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)  # Disabled until results are calculated
        self.export_btn.setMinimumHeight(35)
        results_layout.addWidget(self.export_btn)

        results_group.setLayout(results_layout)
        main_layout.addWidget(results_group)

    def browse_pdf(self):
        """Open file dialog to select a PDF file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select PDF File",
            "",
            "PDF Files (*.pdf);;All Files (*)"
        )

        if file_path:
            self.pdf_path = file_path
            self.file_label.setText(Path(file_path).name)
            self.file_label.setStyleSheet("color: #000; font-weight: bold;")

    def add_table_row(self):
        """Add a new row to the keywords table."""
        row_count = self.keywords_table.rowCount()
        self.keywords_table.insertRow(row_count)

    def remove_table_row(self):
        """Remove the selected row from the keywords table."""
        current_row = self.keywords_table.currentRow()
        if current_row >= 0:
            self.keywords_table.removeRow(current_row)
        else:
            QMessageBox.warning(self, "No Selection", "Please select a row to remove.")

    def clear_table(self):
        """Clear all entries in the keywords table."""
        self.keywords_table.clearContents()

    def get_keywords_from_table(self) -> Dict[str, int]:
        """
        Extract keywords and values from the table.

        Returns:
            Dictionary mapping keywords to their integer values.
        """
        keywords = {}

        for row in range(self.keywords_table.rowCount()):
            keyword_item = self.keywords_table.item(row, 0)
            value_item = self.keywords_table.item(row, 1)

            if keyword_item and value_item:
                keyword = keyword_item.text().strip()
                value_text = value_item.text().strip()

                if keyword and value_text:
                    try:
                        value = int(value_text)
                        keywords[keyword] = value
                    except ValueError:
                        QMessageBox.warning(
                            self,
                            "Invalid Value",
                            f"Value for keyword '{keyword}' must be an integer. Skipping this entry."
                        )

        return keywords

    def extract_text_with_ocr(self, pdf_path: str, keywords: Dict[str, int]):
        """
        Extract text from PDF using OCR for image-based PDFs.
        Runs in a separate thread with progress dialog.

        Args:
            pdf_path: Path to the PDF file.
            keywords: Dictionary of keywords for processing after OCR.
        """
        # Create and configure progress dialog
        self.progress_dialog = QProgressDialog(
            "Initializing OCR...",
            "Cancel",
            0,
            100,
            self
        )
        self.progress_dialog.setWindowTitle("OCR Processing")
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setMinimumDuration(0)
        self.progress_dialog.setValue(0)

        # Create OCR worker thread
        self.ocr_worker = OCRWorker(pdf_path)

        # Connect signals
        self.ocr_worker.progress.connect(self.on_ocr_progress)
        self.ocr_worker.finished.connect(lambda text: self.on_ocr_finished(text, keywords))
        self.ocr_worker.error.connect(self.on_ocr_error)
        self.progress_dialog.canceled.connect(self.on_ocr_canceled)

        # Start OCR processing
        self.ocr_worker.start()

    def on_ocr_progress(self, value: int, message: str):
        """Update progress dialog during OCR processing."""
        if self.progress_dialog:
            self.progress_dialog.setValue(value)
            self.progress_dialog.setLabelText(message)

    def on_ocr_finished(self, text: str, keywords: Dict[str, int]):
        """Handle successful OCR completion."""
        if self.progress_dialog:
            self.progress_dialog.close()

        if not text.strip():
            QMessageBox.warning(
                self,
                "No Text Found",
                "OCR completed but no text was found in the PDF. "
                "The PDF might be empty or the images might not contain readable text."
            )
            return

        # Display extracted text
        self.display_extracted_text(text)

        # Count keywords and display results
        results = self.count_keywords(text, keywords)
        self.display_results(results)

    def on_ocr_error(self, error_message: str):
        """Handle OCR processing errors."""
        if self.progress_dialog:
            self.progress_dialog.close()

        QMessageBox.critical(
            self,
            "OCR Error",
            f"{error_message}\n\n"
            "Please ensure Tesseract OCR is properly installed on your system."
        )

    def on_ocr_canceled(self):
        """Handle OCR cancellation."""
        if self.ocr_worker and self.ocr_worker.isRunning():
            self.ocr_worker.terminate()
            self.ocr_worker.wait()

    def count_keywords(self, text: str, keywords: Dict[str, int]) -> List[Tuple[str, int, int, int]]:
        """
        Count occurrences of keywords in text (case-insensitive).

        Args:
            text: Text to search in.
            keywords: Dictionary mapping keywords to their values.

        Returns:
            List of tuples: (keyword, count, value, subtotal)
        """
        results = []

        for keyword, value in keywords.items():
            # Case-insensitive search using word boundaries
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            count = len(pattern.findall(text))
            subtotal = count * value
            results.append((keyword, count, value, subtotal))

        return results

    def calculate_total(self):
        """Main calculation function triggered by the Calculate button."""
        # Validate PDF file is selected
        if not self.pdf_path:
            QMessageBox.warning(self, "No PDF Selected", "Please select a PDF file first.")
            return

        # Get keywords from table
        keywords = self.get_keywords_from_table()

        if not keywords:
            QMessageBox.warning(
                self,
                "No Keywords",
                "Please enter at least one keyword with a value."
            )
            return

        # Auto-save keywords before processing
        self.save_keywords()

        try:
            # Always use OCR extraction
            self.extract_text_with_ocr(self.pdf_path, keywords)

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def display_extracted_text(self, text: str):
        """
        Display extracted text in the extracted text preview area.

        Args:
            text: The extracted text to display
        """
        # Limit display to first 5000 characters to avoid UI lag
        max_chars = 5000
        display_text = text[:max_chars]

        if len(text) > max_chars:
            display_text += f"\n\n... (showing first {max_chars} characters of {len(text)} total)"

        self.extracted_text.setText(display_text)

    def display_results(self, results: List[Tuple[str, int, int, int]]):
        """
        Display calculation results in the results table.

        Args:
            results: List of tuples containing (keyword, count, value, subtotal)
        """
        # Store results for Excel export
        self.current_results = results

        # Disable sorting while populating
        self.results_table.setSortingEnabled(False)

        # Clear existing data
        self.results_table.setRowCount(0)

        # Calculate grand total
        grand_total = 0
        for keyword, count, value, subtotal in results:
            grand_total += subtotal

        # Add result rows
        sorted_results = sorted(results, key=lambda x: x[0].lower())
        for row_idx, (keyword, count, value, subtotal) in enumerate(sorted_results):
            self.results_table.insertRow(row_idx)

            # Keyword column (left-aligned)
            keyword_item = QTableWidgetItem(keyword)
            self.results_table.setItem(row_idx, 0, keyword_item)

            # Count column (right-aligned)
            count_item = QTableWidgetItem(str(count))
            count_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row_idx, 1, count_item)

            # Value column (right-aligned)
            value_item = QTableWidgetItem(str(value))
            value_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row_idx, 2, value_item)

            # Subtotal column (right-aligned)
            subtotal_item = QTableWidgetItem(str(subtotal))
            subtotal_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.results_table.setItem(row_idx, 3, subtotal_item)

        # Add grand total row
        total_row_idx = len(sorted_results)
        self.results_table.insertRow(total_row_idx)

        # Grand total label (bold)
        total_label = QTableWidgetItem("GRAND TOTAL")
        font = QFont()
        font.setBold(True)
        total_label.setFont(font)
        total_label.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.results_table.setItem(total_row_idx, 0, total_label)

        # Empty cells for count and value
        self.results_table.setItem(total_row_idx, 1, QTableWidgetItem(""))
        self.results_table.setItem(total_row_idx, 2, QTableWidgetItem(""))

        # Grand total value (bold, right-aligned)
        total_value = QTableWidgetItem(str(grand_total))
        total_value.setFont(font)
        total_value.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.results_table.setItem(total_row_idx, 3, total_value)

        # Re-enable sorting
        self.results_table.setSortingEnabled(True)

        # Enable export button
        self.export_btn.setEnabled(True)

    def export_to_excel(self):
        """Export calculation results to Excel file."""
        if not self.current_results:
            QMessageBox.warning(self, "No Results", "No results to export. Please calculate first.")
            return

        # Generate default filename with PDF name
        if self.pdf_path:
            pdf_name = Path(self.pdf_path).stem  # Get filename without extension
            # Sanitize filename (remove special characters)
            pdf_name = "".join(c for c in pdf_name if c.isalnum() or c in (' ', '-', '_')).strip()
            default_filename = f"{pdf_name}_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            default_filename = f"keyword_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            str(Path.home() / default_filename),
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Keyword Results"

            # Add headers
            headers = ["Keyword", "Count", "Value", "Subtotal"]
            ws.append(headers)

            # Style headers (bold)
            header_font = Font(bold=True)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add data rows
            sorted_results = sorted(self.current_results, key=lambda x: x[0].lower())
            grand_total = 0

            for keyword, count, value, subtotal in sorted_results:
                ws.append([keyword, count, value, subtotal])
                grand_total += subtotal

            # Add grand total row
            total_row = len(sorted_results) + 2  # +1 for header, +1 for new row
            ws.append(["GRAND TOTAL", "", "", grand_total])

            # Style grand total row (bold)
            for col_idx in [1, 4]:  # Keyword and Subtotal columns
                cell = ws.cell(row=total_row, column=col_idx)
                cell.font = header_font

            # Adjust column widths
            ws.column_dimensions['A'].width = 30  # Keyword
            ws.column_dimensions['B'].width = 12  # Count
            ws.column_dimensions['C'].width = 12  # Value
            ws.column_dimensions['D'].width = 15  # Subtotal

            # Right-align numeric columns
            for row_idx in range(2, total_row + 1):  # Skip header
                for col_idx in [2, 3, 4]:  # Count, Value, Subtotal
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='right')

            # Save workbook
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Successful",
                f"Results exported successfully to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Could not export to Excel:\n{str(e)}"
            )


def main():
    """Main entry point for the application."""
    app = QApplication(sys.argv)

    # Set application style
    app.setStyle('Fusion')

    window = KeywordEstimatorApp()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
